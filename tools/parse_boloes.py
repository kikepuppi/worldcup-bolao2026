#!/usr/bin/env python3
"""
Parse the friends' World Cup prediction-pool spreadsheets (boloes/*.xlsx, *.xls)
into a single standardized data/palpites.json.

All 16 files share the identical layout: one sheet `BOLAO_2026`.
  Phase 1 (group scores)  : rows 9..80  -> B=nº C=date D=time E=home F=homeGoals G=awayGoals H=away I=group J=city
  Phase 2 (knockout path) : rows 9..56  -> L=group M=team  N..S = cumulative 'x' markers
                            N=Round of 32, O=R16, P=QF, Q=SF, R=Final, S=Champion

.xls files can't be read by openpyxl, so a PowerShell pre-step (tools/convert_xls.ps1)
converts them to xlsx in a temp folder; this parser reads xlsx only.

Usage:  python tools/parse_boloes.py
"""
import json
import os
import glob
import sys
from datetime import date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
BOLOES_DIR = os.path.join(ROOT, "boloes")
CONV_DIR = os.path.join(ROOT, "_conv")          # converted .xls -> .xlsx live here
OUT_PATH = os.path.join(ROOT, "data", "palpites.json")

SHEET = "BOLAO_2026"

# Portuguese sheet name -> the API's exact name_en (worldcup26.ir /get/teams)
PT_TO_EN = {
    "Alemanha": "Germany", "Argentina": "Argentina", "Argélia": "Algeria",
    "Arábia Saudita": "Saudi Arabia", "Austrália": "Australia", "Brasil": "Brazil",
    "Bélgica": "Belgium", "Bósnia": "Bosnia and Herzegovina", "Cabo Verde": "Cape Verde",
    "Canadá": "Canada", "Catar": "Qatar", "Colômbia": "Colombia",
    "Coreia do Sul": "South Korea", "Costa do Marfim": "Ivory Coast", "Croácia": "Croatia",
    "Curaçao": "Curaçao", "Egito": "Egypt", "Equador": "Ecuador", "Escócia": "Scotland",
    "Espanha": "Spain", "Estados Unidos": "United States", "França": "France",
    "Gana": "Ghana", "Haiti": "Haiti", "Holanda": "Netherlands", "Inglaterra": "England",
    "Iraque": "Iraq", "Irã": "Iran", "Japão": "Japan", "Jordânia": "Jordan",
    "Marrocos": "Morocco", "México": "Mexico", "Noruega": "Norway",
    "Nova Zelândia": "New Zealand", "Panamá": "Panama", "Paraguai": "Paraguay",
    "Portugal": "Portugal", "RD Congo": "Democratic Republic of the Congo",
    "República Tcheca": "Czech Republic", "Senegal": "Senegal", "Suécia": "Sweden",
    "Suíça": "Switzerland", "Tunísia": "Tunisia", "Turquia": "Turkey",
    "Uruguai": "Uruguay", "Uzbequistão": "Uzbekistan", "África do Sul": "South Africa",
    "Áustria": "Austria",
}

# the 48 canonical English names (must equal the API team set)
EN_TEAMS = set(PT_TO_EN.values())

# column indices (1-based)
C_NUM, C_HOME, C_HG, C_AG, C_AWAY, C_GROUP = 2, 5, 6, 7, 8, 9
C_TEAM = 13           # M
STAGE_COLS = [14, 15, 16, 17, 18, 19]   # N..S -> stages 1..6


def slug(stem: str) -> str:
    return "".join(c.lower() if c.isalnum() else "_" for c in stem).strip("_")


def to_en(raw, where: str) -> str:
    if raw is None:
        raise ValueError(f"empty team cell at {where}")
    name = str(raw).strip()
    if name not in PT_TO_EN:
        raise ValueError(f"unmapped team name {name!r} at {where}")
    return PT_TO_EN[name]


def as_int(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    return int(round(float(v)))


def parse_workbook(path: str, name: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"{name}: missing sheet {SHEET} (has {wb.sheetnames})")
    ws = wb[SHEET]

    phase1 = []
    for r in range(9, 81):  # 72 matches
        home = ws.cell(r, C_HOME).value
        away = ws.cell(r, C_AWAY).value
        if home is None and away is None:
            continue
        phase1.append({
            "n": as_int(ws.cell(r, C_NUM).value),
            "home": to_en(home, f"{name}!E{r}"),
            "away": to_en(away, f"{name}!H{r}"),
            "group": str(ws.cell(r, C_GROUP).value or "").strip(),
            "ph": as_int(ws.cell(r, C_HG).value),
            "pa": as_int(ws.cell(r, C_AG).value),
        })

    phase2 = {}
    for r in range(9, 57):  # 48 teams
        team_raw = ws.cell(r, C_TEAM).value
        if team_raw is None or not str(team_raw).strip():
            continue
        team = to_en(team_raw, f"{name}!M{r}")
        stage = 0
        for i, col in enumerate(STAGE_COLS, start=1):
            v = ws.cell(r, col).value
            if v is not None and str(v).strip().lower() == "x":
                stage = i
        phase2[team] = stage

    # per-file validation
    if len(phase1) != 72:
        raise ValueError(f"{name}: expected 72 phase-1 matches, got {len(phase1)}")
    if len(phase2) != 48:
        raise ValueError(f"{name}: expected 48 phase-2 teams, got {len(phase2)}")
    missing = EN_TEAMS - set(phase2)
    if missing:
        raise ValueError(f"{name}: phase-2 missing teams {sorted(missing)}")

    return {"id": slug(name), "name": name.replace("_", " "), "phase1": phase1, "phase2": phase2}


def main():
    files = {}  # display-name(stem) -> path  (xlsx wins; converted .xls fill the rest)
    for p in sorted(glob.glob(os.path.join(BOLOES_DIR, "*.xlsx"))):
        files[os.path.splitext(os.path.basename(p))[0]] = p
    for p in sorted(glob.glob(os.path.join(CONV_DIR, "*.xlsx"))):
        files.setdefault(os.path.splitext(os.path.basename(p))[0], p)

    if not files:
        sys.exit("No spreadsheets found. Run tools/convert_xls.ps1 first for the .xls files.")

    participants = []
    for name in sorted(files, key=str.lower):
        participants.append(parse_workbook(files[name], name))
        print(f"  ok  {name:16s} ({files[name]})")

    out = {
        "generated_at": date.today().isoformat(),
        "draw_source": "worldcup26.ir",
        "scoring": {
            "phase1": {"exact": 6, "winner_score": 3, "loser_score": 2, "outcome": 1},
            "phase2": {"r32": 2, "r16": 3, "qf": 5, "sf": 10, "final": 18, "champion": 20},
        },
        "participants": participants,
    }
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"\nWrote {OUT_PATH}: {len(participants)} participants.")


if __name__ == "__main__":
    main()
